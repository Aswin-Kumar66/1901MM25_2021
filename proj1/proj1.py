
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border , Side
import os
import csv

font1 = Font(color='000000FF' , size= '12' , name= 'Century')       #blue for  correct options 
font2 = Font(color='FF0000' , size= '12' , name= 'Century')         #red colour for incorrect answer
font3 = Font(color='008000' , size='12' ,name= 'Century')         #green colour for correct answer
os.mkdir('./output')
path = os.getcwd()
os.chdir('./sample_input')
with open('master_roll.csv' , 'r') as f:
 names = list(csv.reader(f))
 with open('responses.csv' , 'r') as g :
     responses  = list(csv.reader(g))

for roll in names[1:]:
    os.chdir(path)
    os.chdir('./output')
    wb= Workbook()
    wb.create_sheet(index = 0 ,title  = 'quiz')
    correct_answers = 0
    incorrect_answers = 0
    not_attempted = 0 
    quiz = wb['quiz']
    quiz['A6'] ,quiz['A7'] , quiz['D6'] , quiz['E6'] = 'Name:' , 'Roll Number:' , 'Exam:' , 'quiz'
    quiz['A10'] ,quiz['A11'] , quiz['A12'] = 'No.' , 'Marking:' , 'Total'
    quiz['A10'].font ,quiz['A11'].font , quiz['A12'].font = Font(bold = True,size='12' ,name= 'Century'),Font(bold = True,size='12' ,name= 'Century'),Font(bold = True,size='12' , name= 'Century')
    quiz['A15'] , quiz['B15'] = 'Student Ans' , 'Correct Ans'
    quiz['D15'] , quiz['E15'] = 'Student Ans' , 'Correct Ans'
    quiz['A15'].font , quiz['B15'].font =  Font(bold = True , size='12' , name= 'Century'),Font(bold = True ,size='12' , name= 'Century')
    quiz['D15'].font , quiz['E15'].font =  Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century')
    quiz['B11'],quiz['C11'] , quiz['D11']= 5 , -1 , 0
    quiz['B11'].font,quiz['C11'].font  = font3 , font2
    quiz['B9'],quiz['C9'],quiz['D9'],quiz['E9'] = 'Right' , 'Wrong' , 'Not Attempt' , 'Max'
    quiz['B9'].font,quiz['C9'].font,quiz['D9'].font,quiz['E9'].font = Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century') , Font(bold = True,size='12' , name= 'Century')
    quiz['A5'] = 'Mark Sheet'
    quiz['A5'].font = Font( underline= 'single' , size='18')
    quiz['B16'] , quiz['B17'] , quiz['B18'],quiz['B19']  = 'Option A' , 'Option D' , 'Option B' , 'Option C'
    quiz['B20'] , quiz['B21'],quiz['B22'] , quiz['B23']  = 'Option B' , 'Option C' , 'Option D' , 'Option D'
    quiz['B24'] , quiz['B25'] , quiz['B26'],quiz['B27']  = 'Option A' , 'Option A' , 'Option C' , 'Option A'
    quiz['B28'] , quiz['B29'],quiz['B30'] , quiz['B31']  = 'Option D' , 'Option D' , 'Option B' , 'Option D'
    quiz['B32'] , quiz['B33'] , quiz['B34'],quiz['B35']  = 'Option C' , 'Option D' , 'Option B' , 'Option D'
    quiz['B36'] , quiz['B37'],quiz['B38'] , quiz['B39'] , quiz['B40']  = 'Option A' , 'Option A' , 'Option A' , 'Option D' , 'Option D'
    quiz['E16'] , quiz['E17'] , quiz['E18']  = 'Option A' , 'Option C' , 'Option D' 
    quiz['E16'].font , quiz['E17'].font , quiz['E18'].font  = font1 , font1 , font1
    for i in range(16,41) :
        string = 'B' + str(i)
        quiz[string].font = font1 
    for response in responses :
        if(response[3] == roll[1]) :
              quiz['A16'] , quiz['A17'] , quiz['A18'],quiz['A19']  = response[7] , response[8] , response[9] ,response[10]
              quiz['A20'] , quiz['A21'],quiz['A22'] , quiz['A23']  = response[11] , response[12] , response[13] ,response[14]
              quiz['A24'] , quiz['A25'] , quiz['A26'],quiz['A27']  = response[15] , response[16] , response[17] ,response[18]
              quiz['A28'] , quiz['A29'],quiz['A30'] , quiz['A31']  = response[19] , response[20] , response[21] ,response[22]
              quiz['A32'] , quiz['A33'] , quiz['A34'],quiz['A35']  = response[23] , response[24] , response[25] ,response[26]
              quiz['A36'] , quiz['A37'],quiz['A38'] , quiz['A39'] , quiz['A40']  = response[27] , response[28] , response[29] ,response[30],response[31]
              quiz['D16'] , quiz['D17'] , quiz['D18']  = response[32] , response[33] , response[34] 
    for i in range(16,41) :
        string1 = 'A' + str(i)
        string2 = 'B' + str(i)
        if(quiz[string1].value == quiz[string2].value) :
            quiz[string1].font = font3
            correct_answers = correct_answers + 1
        elif (quiz[string1].value != quiz[string2].value) and quiz[string1].value != '' :   
            quiz[string1].font = font2     
            incorrect_answers = incorrect_answers + 1
        else :
            not_attempted = not_attempted + 1
    for i in range(16,19) :
        string1 = 'D' + str(i)
        string2 = 'E' + str(i) 
        if(quiz[string1].value == quiz[string2].value) :
            quiz[string1].font = font3
            correct_answers = correct_answers + 1
        elif (quiz[string1].value != quiz[string2].value) and quiz[string1].value != '' :   
            quiz[string1].font = font2     
            incorrect_answers = incorrect_answers + 1
        else :
            not_attempted = not_attempted + 1        
    quiz['B10'] ,quiz['C10'] , quiz['D10'] ,quiz['E10'] = correct_answers , incorrect_answers , not_attempted , 28
    quiz['B12'] ,quiz['C12']  = (quiz['B10'].value * quiz['B11'].value),(quiz['C10'].value * quiz['C11'].value)
    quiz['E12'] = quiz['B12'].value + quiz['C12'].value
    quiz['E12'] = str(quiz['E12'].value) + '/140'
    quiz['E12'].font = font1
    quiz['B12'].font , quiz['B10'].font = font3 , font3
    quiz['C12'].font , quiz['C10'].font = font2 , font2
    quiz['B7'] , quiz['B6'] = roll[0] , roll[1]
    quiz['B6'].font,quiz['B7'].font,quiz['E6'].font = Font(bold = True , size='12' , name= 'Century'),Font(bold = True , size='12' , name= 'Century'),Font(bold = True , size='12' , name= 'Century') 
    quiz.column_dimensions['A'].width = 16.89
    quiz.column_dimensions['B'].width = 16.89
    quiz.column_dimensions['C'].width = 16.89
    quiz.column_dimensions['D'].width = 16.89
    quiz.column_dimensions['E'].width = 16.89
    quiz.merge_cells('A5:E5')
    thin_border = Border(left = Side(style='thin') ,right = Side(style='thin') ,top = Side(style='thin') ,bottom = Side(style='thin')  )
    for i in range(15,41) :
     quiz.cell(row = i , column = 1).border = thin_border
     quiz.cell(row = i , column = 2).border = thin_border
    for i in range(9,13) :
     quiz.cell(row = i , column = 1).border = thin_border
     quiz.cell(row = i , column = 2).border = thin_border
     quiz.cell(row = i , column = 3).border = thin_border
     quiz.cell(row = i , column = 4).border = thin_border
     quiz.cell(row = i , column = 5).border = thin_border
    for i in range(15,19) :
     quiz.cell(row = i , column = 4).border = thin_border
     quiz.cell(row = i , column = 5).border = thin_border
    #quiz.column_dimensions['B'].width = 25  
    for i in range(quiz.max_column) :
     for row in quiz[2:quiz.max_row]:
       cell = row[i]
       cell.alignment = Alignment(horizontal='center') 
    img = openpyxl.drawing.image.Image( path + '/iitp logo.png')
    img.anchor = 'A1'
    img.width = 600
    img.height = 85
    quiz.add_image(img)
    wb.save(roll[0] + '.xlsx')

        

  


