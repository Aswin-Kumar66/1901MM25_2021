from openpyxl import Workbook
import os
import csv
from openpyxl import load_workbook
def convert(s) :
    if(s == 'AA') :
        return 10
    elif s == "AB" :
        return 9
    elif s == 'BB' :
        return 8
    elif s == 'BC' :
        return 7
    elif s == 'CC' :
        return 6
    elif s == 'CD' :
        return 5
    elif s == 'DD' :
        return 4
    elif s == 'I' :
        return 0
    elif s =="F" :
        return 0

def generate_marksheet():
    os.mkdir('output')
    head = ['Sl. No','Subject No' , 'Subject Name' , 'L-T-P' , 'Credit' , 'Subject Type' , 'Grade']

    with open('grades.csv' , 'r') as f :
     grades = list(csv.reader(f))
     with open('subjects_master.csv', 'r') as g :
      subjects = list(csv.reader(g) )
      with open('names-roll.csv' , 'r') as h :
          names  = list(csv.reader(h))
          sub_dict = {}
    for course in subjects:
        if course[0] not in sub_dict:
            sub_dict[course[0]] = course

    for roll in names[1:]:
        j=1
        wb= Workbook()
        wb.create_sheet(index = 0 ,title  = 'Overall')
        overall = wb['Overall']
        overall['A1'] ,overall['B1'] = 'Roll No .' , roll[0]
        overall['A2'] , overall['B2']= 'Name of student ' , roll[1]
        overall['A3'] , overall['B3']= 'Discipline ' , roll[0][4] + roll[0][5]
        overall['A4'], overall['B4'], overall['C4'], overall['D4'], overall['E4'] , overall['F4'] ,  overall['G4'] ,  overall['H4'] , overall['I4']= 'Semester No .' ,1 ,2,3,4,5,6,7,8
        overall['A5'] = 'Semester wise credits taken '
        overall['A6'] = 'SPI '
        overall['A7'] = 'Total Credits '
        overall['A8'] = 'CPI'

        while j<=8 :
            if f'Sem{j}' not in wb.sheetnames:
                wb.create_sheet(f'Sem{j}',j)         
            ws = wb[f"Sem{j}"]
           
            ws.append(head)
            i=1
            for x in grades[1:]:
                
                if x[0]==roll[0] and int(x[1])==j :
                    
                    data=[]
                    data.append(i)
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    ws.append(data)
                    i+=1
            j+=1
            wb.save(f'output\\{roll[0]}.xlsx')
    with open('grades.csv' , 'r') as f :
        grades = csv.reader(f) 
        os.chdir(r'output')
        for line in grades  :  
            if line[0] != 'Roll'   :  
                 sheetName = line[0] + '.xlsx'   
                
                 wb = load_workbook(sheetName)
                 credits1 , spi1 ,total ,sum = 0 , 0 , 0 , 0 
                 credits2 , spi2 = 0 ,0
                 credits3 , spi3 = 0 , 0
                 credits4  , spi4= 0 , 0
                 credits5  , spi5= 0 ,0
                 credits6  , spi6= 0 ,0
                 credits7  , spi7= 0 , 0
                 credits8  , spi8= 0 , 0
                
                 sheet1 = wb['Sem1']
                 sheet2 = wb['Sem2']
                 sheet3 = wb['Sem3']
                 sheet4 = wb['Sem4']
                 sheet5 = wb['Sem5']
                 sheet6 = wb['Sem6']
                 sheet7 = wb['Sem7']
                 sheet8 = wb['Sem8']   
                 overall = wb['Overall']    
                 maxrow1  = sheet1.max_row
                 maxrow2  = sheet2.max_row
                 maxrow3  = sheet3.max_row
                 maxrow4  = sheet4.max_row
                 maxrow5  = sheet5.max_row
                 maxrow6  = sheet6.max_row
                 maxrow7  = sheet7.max_row
                 maxrow8  = sheet8.max_row
                 for i in range(maxrow1-1) :
                    credits1 =  credits1 +  sheet1.cell(row=i + 2, column=5).value
                    spi1  = spi1 + (sheet1.cell(row=i + 2, column=5).value)*convert((sheet1.cell(row=i + 2, column=7).value)) 
                         
                 for i in range(maxrow2-1) :
                       credits2 =  credits2 +  sheet2.cell(row=i + 2, column=5).value
                       spi2  = spi2 + (sheet2.cell(row=i + 2, column=5).value)*convert((sheet2.cell(row=i + 2, column=7).value))
                 for i in range(maxrow3-1) :
                       credits3 =  credits3 +  sheet3.cell(row=i + 2, column=5).value
                       spi3  = spi3 + (sheet3.cell(row=i + 2, column=5).value)*convert((sheet3.cell(row=i + 2, column=7).value))
                 for i in range(maxrow4-1) :
                         credits4 =  credits4 +  sheet4.cell(row=i + 2, column=5).value
                         spi4  = spi4 + (sheet4.cell(row=i + 2, column=5).value)*convert((sheet4.cell(row=i + 2, column=7).value))
                 for i in range(maxrow5-1) :
                        credits5 =  credits5 +  sheet5.cell(row=i + 2, column=5).value
                        spi5  = spi5 + (sheet5.cell(row=i + 2, column=5).value)*convert((sheet5.cell(row=i + 2, column=7).value))
                 for i in range(maxrow6-1) :
                   credits6 =  credits6 +  sheet6.cell(row=i + 2, column=5).value
                   spi6  = spi6 + (sheet6.cell(row=i + 2, column=5).value)*convert((sheet6.cell(row=i + 2, column=7).value))
                 for i in range(maxrow7-1) :
                      credits7 =  credits7 +  sheet7.cell(row=i + 2, column=5).value
                      spi7  = spi7 + (sheet7.cell(row=i + 2, column=5).value)*convert((sheet7.cell(row=i + 2, column=7).value))
                 for i in range(maxrow8-1) :
                         credits8 =  credits8 +  sheet8.cell(row=i + 2, column=5).value
                         spi8  = spi8 + (sheet8.cell(row=i + 2, column=5).value)*convert((sheet8.cell(row=i + 2, column=7).value))
                
                 overall['B5'] , overall['C5'] , overall['D5'] , overall['E5']  = credits1 , credits2 , credits3 , credits4  
                 overall['F5'] , overall['G5'] , overall['H5'] , overall['I5']  = credits5 , credits6 , credits7 , credits8  
                 overall['B6'] , overall['C6'] , overall['D6'] , overall['E6']= round(spi1/credits1 ,2)  , round(spi2/credits2, 2) ,round(spi3/credits3,2) , round(spi4/credits4,2)  
                 overall['F6'] , overall['G6'] , overall['H6'] , overall['I6']  = round(spi5/credits5,2) , round(spi6/credits6,2) , round(spi7/credits7,2) , round(spi8/credits8,2)  
        
                 for i in range(8)     :
                  total = total + overall.cell(row = 5 , column =i + 2).value
                  sum = sum + (overall.cell(row = 6 , column = i+2).value)*(overall.cell(row = 5 , column = i+2).value)
                  overall.cell(row = 8 ,column = i+2).value = round(sum/total , 2)
                  overall.cell(row=7 , column = i + 2).value = total
                 overall.column_dimensions['A'].width = 25
                 sheet1.column_dimensions['C'].width = 25
        
                 wb.save(sheetName)
        
                 
    return 

generate_marksheet()
