'''import os
from fpdf import FPDF
import csv

def generate_transcript():
  os.chdir('sample_input')
  head = ['Subj.code' , 'Subject Name' , 'L-T-P' , 'Credit' , 'Subject Type' , 'Grade']
  with open('grades.csv' , 'r') as f :
     grades = list(csv.reader(f))
     with open('subjects_master.csv', 'r') as g :
      subjects = list(csv.reader(g) )
      with open('names-roll.csv' , 'r') as h :
          names  = list(csv.reader(h))
          sub_dict = {}
  for course in subjects:
        if course[0] not in sub_dict:
            sub_dict[course[0]] = course
  for roll in names[1:]:
      
      sem1 ,sem2 , sem3 ,  sem4 , sem5 , sem6 , sem7 , sem8= [],[],[],[],[],[],[],[]
      sem1.append(head) , sem2.append(head) ,sem3.append(head) ,sem4.append(head) ,sem5.append(head) ,sem6.append(head) ,sem7.append(head) ,sem8.append(head) 
      
      
      for x in grades[1:]:
            if x[0]==roll[0] and int(x[1])==1 :
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem1.append(data)
            elif x[0]==roll[0] and int(x[1])==2 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem2.append(data)     
            elif x[0]==roll[0] and int(x[1])==3 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem3.append(data)  
            elif x[0]==roll[0] and int(x[1])==3 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem3.append(data)  
            elif x[0]==roll[0] and int(x[1])==4 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem4.append(data)
            elif x[0]==roll[0] and int(x[1])==5 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem5.append(data)  
            elif x[0]==roll[0] and int(x[1])==6 :  
                    data=[]
                    
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    sem6.append(data)    
            









  return
   


generate_transcript()
'''
from openpyxl import Workbook
import os
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
def generate_marksheet():
    os.mkdir('output')
    path = os.getcwd()
    os.chdir('sample_input')
    head = ['Sl. No','Subject No' , 'Subject Name' , 'L-T-P' , 'Credit' , 'Subject Type' , 'Grade']

    with open('grades.csv' , 'r') as f :
     grades = list(csv.reader(f))
     with open('subjects_master.csv', 'r') as g :
      subjects = list(csv.reader(g) )
      with open('names-roll.csv' , 'r') as h :
          names  = list(csv.reader(h))
          sub_dict = {}
    for course in subjects:
        if course[0] not in sub_dict:
            sub_dict[course[0]] = course

    for roll in names[1:]:
        j=1
        wb= Workbook()
        wb.create_sheet(index = 0 ,title  = 'Overall')
        overall = wb['Overall']
        overall['A1'] ,overall['B1'] = 'Roll No .' , roll[0]
        overall['A2'] , overall['B2']= 'Name of student ' , roll[1]
        overall['A3'] , overall['B3']= 'Discipline ' , roll[0][4] + roll[0][5]
        overall['A4'], overall['B4'], overall['C4'], overall['D4'], overall['E4'] , overall['F4'] ,  overall['G4'] ,  overall['H4'] , overall['I4']= 'Semester No .' ,1 ,2,3,4,5,6,7,8
        overall['A5'] = 'Semester wise credits taken '
        overall['A6'] = 'SPI '
        overall['A7'] = 'Total Credits '
        overall['A8'] = 'CPI'

        while j<=8 :
            if f'Sem{j}' not in wb.sheetnames:
                wb.create_sheet(f'Sem{j}',j)         
            ws = wb[f"Sem{j}"]
           
            ws.append(head)
            i=1
            for x in grades[1:]:
                
                if x[0]==roll[0] and int(x[1])==j :
                    
                    data=[]
                    data.append(i)
                    data.append(x[2])
                    y=sub_dict[x[2]]               
                    data.append(y[1])
                    data.append(y[2])
                    data.append(int(y[3]))
                    data.append(x[5])
                    data.append(x[4])
                    ws.append(data)
                    i+=1
            j+=1
            os.chdir(path)
            os.chdir('output')
            wb.save(f'{roll[0]}.pdf')

generate_marksheet()